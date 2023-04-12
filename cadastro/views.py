from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, HttpResponseNotFound, HttpResponseRedirect
from django.urls import reverse
from django.core.exceptions import ObjectDoesNotExist
import os
from .forms import CadastroForm, \
    Pesquisaform, \
    Dataform, \
    Pesquisaform_numero, \
    LoginForm, \
    Cadastro_imagem_Form

from .models import Cadastro, Imagem
from django.db.models import Q
from django.contrib.auth import authenticate, \
    login as auth_login, \
    logout as auth_logout

from openpyxl import load_workbook
import os


def teste(request):
    return HttpResponse('olá galera')


@login_required
def home(request):
    username = ''
    if request.user.is_authenticated:
        username = request.user
    return render(request, 'home.html', {'username': username})


# Inverte a data para Ano - Mes - dia
def corrigi_data(data):

    dia = data[:2]
    mes = data[3:5]
    ano = data[6:10]

    return ano+'-'+mes+'-'+dia


# Salva no banco o formulário se for válido
def salvar_formulario(formulario, dicionario, CadastroForm):
    if formulario.is_valid():

        formulario.save()
        dicionario['susesso'] = True

        formulario = CadastroForm()
    else:
        dicionario['susesso'] = False
        dicionario['campo'], dicionario['mensagem'] = formulario.errors.as_data().popitem()


# Salva o formulário no Banco
@login_required
def cadastro(request):
    dicionario = {}
    dicionario['campo'] = None
    dicionario['mensagem'] = None
    dicionario['editar'] = False
    dicionario['susesso'] = None
    dicionario['salvar'] = True
    dicionario['editar'] = False
    dicionario['id'] = 0
    if request.method == 'POST':

        data_do_processo = request.POST.get('data_do_processo')

        images = request.FILES.get('images')

        print(images)

        post_data = request.POST.copy()

        post_data['data_do_processo'] = corrigi_data(data_do_processo)

        data_do_recebimento = request.POST.get('data_do_recebimento')

        post_data['data_do_recebimento'] = corrigi_data(data_do_recebimento)

        formulario = CadastroForm(post_data)

        salvar_formulario(formulario, dicionario, CadastroForm)

        formulario = CadastroForm()

    else:
        formulario = CadastroForm()
    dicionario['form'] = formulario
    dicionario['username'] = request.user

    return render(request, 'cadastro.html', dicionario)


@login_required
def pesquisa(request):
    return render(request, 'base_pesquisa.html')


# Pesquisa por data de inicio e fim
@login_required
def novapesquisa(request):
    if request.method == 'GET':
        form = Dataform
        context = {'form': form}
        return render(request, 'base_pesquisa.html', context)

    if (request.method == 'POST'):

        form = Dataform(request.POST)

        if form.is_valid():

            data_inicial = form.data['data_inicio']
            data_final = form.data['data_fim']

            results = Cadastro.objects.filter(Q(
                data_do_recebimento__gte=corrigi_data(data_inicial)) & Q(
                data_do_recebimento__lte=corrigi_data(data_final)))

            context = {'form': form, 'results': results}
            context['username'] = request.user
            return render(request, 'base_pesquisa.html', context)


# Pesquisa por nome
@login_required
def pesquisa_string(request):
    if request.method == 'GET':

        form = Pesquisaform
        context = {'form': form}
        return render(request, 'base_pesquisa.html', context)

    if (request.method == 'POST'):

        form = Pesquisaform(request.POST)
        query = form.data['query']

        if form.is_valid():

            results = Cadastro.objects.filter(requerente__icontains=query)

            context = {'form': form, 'results': results}
            context['username'] = request.user
            return render(request, 'base_pesquisa.html', context)
    else:
        results = Cadastro.objects.all

        context = {'form': form, 'results': results}
        context['username'] = request.user
        return render(request, 'base_pesquisa.html', context)


# Pesquisa por número de processo
@login_required
def pesquisa_numero(request):
    if request.method == 'GET':

        form = Pesquisaform_numero
        context = {'form': form}
        return render(request, 'base_pesquisa.html', context)

    if (request.method == 'POST'):

        form = Pesquisaform(request.POST)
        query = form.data['query']

        if form.is_valid():

            results = Cadastro.objects.filter(
                numero_do_processo__icontains=query)

            context = {'form': form, 'results': results}
            context['username'] = request.user
            return render(request, 'base_pesquisa.html', context)


@login_required
def editar_cadastro(request, variavel):
    dicionario = {}
    dicionario['susesso'] = None
    dicionario['editar'] = True
    dicionario['salvar'] = True

    try:
        objeto = Cadastro.objects.get(id=variavel)
    except ObjectDoesNotExist:
        return render(request, 'cadastro.html', {'susesso':False, 'mensagem':'Cadastro não encontrado'})
    else:

        form = CadastroForm(instance=objeto)

        dicionario['form'] = form

        dicionario['id'] = variavel

        if request.method == 'GET':

            return render(request, 'cadastro.html', dicionario)
        else:

            data_do_processo = request.POST.get('data_do_processo')

            post_data = request.POST.copy()

            post_data['data_do_processo'] = corrigi_data(data_do_processo)

            data_do_recebimento = request.POST.get('data_do_recebimento')

            post_data['data_do_recebimento'] = corrigi_data(
                data_do_recebimento)

            formulario = CadastroForm(post_data, instance=objeto)

            salvar_formulario(formulario, dicionario, CadastroForm)

            dicionario['form'] = CadastroForm()

            dicionario['username'] = request.user
            return render(request, 'cadastro.html', dicionario)


@login_required
def excluir(request, variavel):

    try:
        objeto = Cadastro.objects.get(id=variavel)
    except ObjectDoesNotExist:
        return render(request, 'cadastro.html', {'susesso': False, 'mensagem': 'Cadastro não encontrado'})
    else:

        objeto.delete()
        return render(request, 'cadastro.html', {'alerta': True})


def log_in(request):
    if request.method == 'POST':
        form = LoginForm(request.POST)
        username = request.POST['username']
        password = request.POST['password']

        user = authenticate(username=username, password=password)

        if request.user.is_authenticated:
            return redirect('/registro/home')
        else:
            sucesso = None
            if user is not None:
                sucesso = False
                auth_login(request, user)
                return redirect('/registro/home')
            else:
                sucesso = True
                # Exibir uma mensagem de erro se a autenticação falhar
                return render(request, 'login.html',
                              {'form': form, 'sucesso': sucesso})
    else:
        # Se for um pedido GET, exibir o formulário de login vazio
        form = LoginForm()
    return render(request, 'login.html', {'form': form})

@login_required
def logout(request):
    auth_logout(request)
    return redirect('/registro/login')


@login_required
def upload(request):
    result = None
    if (request.method == 'GET'):
        return render(request, 'upload.html')
    
    elif request.method == 'POST':

        # Obtém o arquivo enviado através da solicitação POST
        try:
            file = request.FILES['xlsx_file']
        except:
            result = False
            return render(request, 'upload.html', {'result': result})
        else:
            # Verifica se 
            try:
                wb = load_workbook(filename=file)
            except:
                result = False
                return render(request, 'upload.html', {'result': result})
            else:
                # Selecionar a primeira planilha
                ws = wb.active

                # Ler a primeira linha
                first_row = [cell.value for cell in ws[1]]

                # Verificar se a primeira célula contém o nome "requerente"
                if first_row[0] == 'Requerente ':
                    result = True
                    valid_extensions = ['.xls', '.xlsx']
                    ext = os.path.splitext(file.name)[1]

                    if not ext.lower() in valid_extensions:
                        result = False
                    else:
                        result = True

                        # Carrega o arquivo com a biblioteca openpyxl
                        workbook = load_workbook(file)

                        # Acesse as folhas do arquivo e faça algo com os dados
                        sheet = workbook.active

                        dicionario = {}

                        for row in sheet.iter_rows(values_only=True):
                            if (row[0] == 'Requerente '):
                                continue

                            formulario = Cadastro(
                                requerente=row[0],
                                assunto=row[1],
                                numero_do_processo=row[2],
                                ano=row[3],
                                data_do_processo=row[4],
                                data_do_recebimento=row[5],
                                responsavel=row[6],
                                status=row[7],
                                destino=row[8])

                            formulario.save()

                        # Renderiza uma resposta com uma mensagem de sucesso
                        return render(request, 'upload.html', {'result': result})
                else:
                    result = False
                    return render(request, 'upload.html', {'result': result})

            return render(request, 'upload.html', {'result': result})

@login_required
def anexo(request, id):

    dicionario = {}
    dicionario['id']=id
    dicionario['sucesso'] = None
    form = Cadastro.objects.get(pk=id)
    form2 = Cadastro_imagem_Form(instance=form)
    dicionario['form'] = form2
    dicionario['cadastro'] = [form]
    if request.method == 'POST':
        fotos = request.FILES.getlist('imagens')
        pdf = request.FILES.getlist('pdf')
        for img in fotos:
            print("entrei em imagens")
            imagem = Imagem.objects.create(cadastro=form, imagem=img, pdf=None)
            dicionario['sucesso'] = True

        for arq in pdf:
            print("entrei em PDF")
            arquivo = Imagem.objects.create(cadastro=form, pdf=arq, imagem=None)
            dicionario['sucesso'] = True

        return render(request, 'anexo.html',dicionario)
    elif request.method == 'GET':
        return render(request, 'anexo.html', dicionario)

# DELETA AS IMAGENS MARCADAS NO CHECKBOX
def delete_images(request, id):
    # SALVA URL DA VIEW ANEXO ACRESCENTANDO O ID
    url = reverse('anexo', kwargs={'id': id})
    if request.method == 'POST':
        # Obter a lista de IDs de imagem selecionados
        imagens_ids = request.POST.getlist('imagens[]')
        print(id)
        print(imagens_ids)
        cadastro = get_object_or_404(Cadastro, pk=id)
        for i in imagens_ids:
            imagem = get_object_or_404(Imagem, pk=i, cadastro=cadastro)
            print(imagem.imagem)
            try:
                temp = imagem.imagem.path
            except:
                temp = imagem.pdf.path
            imagem.delete()
            os.remove(temp)

        return HttpResponseRedirect(url)

    return HttpResponseRedirect(url)

def teste(request):
    if request.method == 'GET':
        return render(request, 'teste.html')

